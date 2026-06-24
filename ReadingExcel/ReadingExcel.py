import openpyxl

# File ---> Workbook ---> Sheets ---> Rows ---> Cells

file=(r"C:\Users\santo\Downloads\May.xlsx")
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row #counts no. of rows in excel sheet 8
cols=sheet.max_column #counts no. of coloumns in excel sheet 6

#Reading all the rows and coloumns from excel sheet

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value, end='          ')
    print()

